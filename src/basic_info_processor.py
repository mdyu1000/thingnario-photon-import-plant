from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
import requests
import logging
import os
from src.constants.constants import SheetNames


class BasicInfoProcessor:
    def __init__(self):
        self.api_key = os.getenv("GOOGLE_MAPS_API_KEY")

    def setup_basic_info_sheet(self, workbook: Workbook) -> Worksheet:
        """
        設置基本資訊分頁的格式和結構
        """
        if SheetNames.BASIC_INFO.value in workbook.sheetnames:
            sheet = workbook[SheetNames.BASIC_INFO.value]
        else:
            sheet = workbook.create_sheet(SheetNames.BASIC_INFO.value)

        # 刪除默認的 Sheet
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]

        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 40
        sheet.column_dimensions["C"].width = 40

        # 設置標題行
        headers = [
            ("電站資訊", "A1"),
            ("答案", "B1"),
        ]

        for header_text, cell in headers:
            sheet[cell] = header_text
            sheet[cell].font = Font(bold=True, color="FFFFFF")
            sheet[cell].fill = PatternFill(
                start_color="1C4587", end_color="1C4587", fill_type="solid"
            )

        # 設置基本資訊欄位
        info_fields = [
            ("地區", "A2"),
            ("電站代碼", "A3"),
            ("電站名稱", "A4"),
            ("地址", "A5"),
            ("緯度", "A6"),
            ("經度", "A7"),
            ("併聯日期", "A8"),
            ("電力結構", "A9"),
            ("註冊碼", "A10"),
        ]

        for field, cell in info_fields:
            sheet[cell] = field

        return sheet

    def fill_station_info(self, sheet: Worksheet, csv_row: dict):
        """
        填充電站基本資訊到指定的工作表
        """
        # 填充動態資料
        sheet["B2"] = csv_row["地區"]
        sheet["B3"] = csv_row["電站代碼"]
        sheet["B4"] = csv_row["電站名稱"]
        sheet["B5"] = csv_row["地址"]

        # 處理經緯度
        full_address = f"{csv_row['地區']}{csv_row['CITY']}{csv_row['地址']}"
        lat, lng = self.get_coordinates_from_google(full_address)
        if lat and lng:
            sheet["B6"] = lat
            sheet["B7"] = lng

        # 設置固定值
        sheet["B9"] = "純光電"
        sheet["B10"] = csv_row["註冊碼"]

    def get_coordinates_from_google(self, address):
        """
        從 Google Maps API 獲取地址的經緯度
        """
        try:
            base_url = "https://maps.googleapis.com/maps/api/geocode/json"
            params = {"address": address, "key": self.api_key}
            response = requests.get(base_url, params=params)
            data = response.json()

            if data["status"] == "OK":
                location = data["results"][0]["geometry"]["location"]
                return location["lat"], location["lng"]
            else:
                logging.error(f"無法獲取地址'{address}'的經緯度: {data['status']}")
                return None, None
        except Exception as e:
            logging.error(f"調用 Google Maps API 時發生錯誤: {str(e)}")
            return None, None
