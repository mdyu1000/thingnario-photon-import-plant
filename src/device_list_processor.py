from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from src.constants.constants import SheetNames
from src.constants.sheet_headers import SheetHeaders
from src.constants.sheet_examples import SheetExamples
from openpyxl.styles import Font, PatternFill
from src.plant_provider import Plant
from typing import List, Tuple


class DeviceListProcessor:
    def __init__(self, plant: Plant):
        self.plant = plant
        self.current_row = 1
        self.current_modbus_id = 1

    def setup_device_list_sheet(self, workbook: Workbook) -> Worksheet:
        if SheetNames.DEVICE_LIST.value in workbook.sheetnames:
            sheet = workbook[SheetNames.DEVICE_LIST.value]
        else:
            sheet = workbook.create_sheet(SheetNames.DEVICE_LIST.value)

        sheet.sheet_format.defaultColWidth = 15
        sheet.column_dimensions["B"].width = 30
        sheet.column_dimensions["C"].width = 30

        return sheet

    def _fill_headers(self, sheet: Worksheet, headers: List[Tuple[str, str]]):
        """直接填充表頭"""
        for header_text, col in headers:
            cell = f"{col}{self.current_row}"
            sheet[cell] = header_text
            sheet[cell].font = Font(bold=True, color="FFFFFF")
            sheet[cell].fill = PatternFill(
                start_color="38761C", end_color="38761C", fill_type="solid"
            )
        self.current_row += 1

    def _fill_example(self, sheet: Worksheet, examples: List[Tuple[str, str]]):
        """直接填充範例"""
        for example_text, col in examples:
            cell = f"{col}{self.current_row}"
            sheet[cell] = example_text
            sheet[cell].font = Font(color="000000")
            sheet[cell].fill = PatternFill(
                start_color="F3F3F3", end_color="F3F3F3", fill_type="solid"
            )
        self.current_row += 1

    def fill_logger(self, sheet: Worksheet, row: dict):
        self.current_row = 1
        self._fill_headers(sheet, SheetHeaders.LOGGER_HEADERS)
        self._fill_example(sheet, SheetExamples.LOGGER_EXAMPLE)

        sheet["B3"] = row["假MAC"]
        sheet["C3"] = "PT-2020"
        self.current_row += 1
        self.current_row += 1

    def fill_pyranometer(self, sheet: Worksheet):
        """填充日照計資訊"""
        self._fill_headers(sheet, SheetHeaders.PYRANOMETER_HEADERS)
        self._fill_example(sheet, SheetExamples.PYRANOMETER_EXAMPLE)

        for pyranometer in self.plant.pyranometers:
            sheet[f"B{self.current_row}"] = pyranometer.device_id
            sheet[f"C{self.current_row}"] = "ADTEK_CS1"
            sheet[f"D{self.current_row}"] = "2"
            sheet[f"E{self.current_row}"] = "9600,N,8,1"
            sheet[f"H{self.current_row}"] = self.current_modbus_id
            self.current_modbus_id += 1
            self.current_row += 1

        self.current_row += 1

    def fill_thermometer(self, sheet: Worksheet):
        self._fill_headers(sheet, SheetHeaders.THERMOMETER_HEADERS)
        self._fill_example(sheet, SheetExamples.THERMOMETER_EXAMPLE)

        for thermometer in self.plant.thermometers:
            sheet[f"B{self.current_row}"] = thermometer.device_id
            sheet[f"C{self.current_row}"] = "ADTEK_CS1"
            sheet[f"D{self.current_row}"] = "2"
            sheet[f"E{self.current_row}"] = "9600, N, 8, 1"
            sheet[f"H{self.current_row}"] = self.current_modbus_id
            self.current_modbus_id += 1
            self.current_row += 1
        self.current_row += 1

    def fill_unused_device(self, sheet: Worksheet):
        self._fill_headers(sheet, SheetHeaders.ANEMOMETER_HEADERS)
        self._fill_example(sheet, SheetExamples.ANEMOMETER_EXAMPLE)
        self.current_row += 1

        self._fill_headers(sheet, SheetHeaders.POWER_METER_HEADERS)
        self._fill_example(sheet, SheetExamples.POWER_METER_EXAMPLE)
        self.current_row += 1

        self._fill_headers(sheet, SheetHeaders.PROTECTION_RELAY_HEADERS)
        self._fill_example(sheet, SheetExamples.PROTECTION_RELAY_EXAMPLE)
        self.current_row += 1

    def fill_inverter(self, sheet: Worksheet):
        self._fill_headers(sheet, SheetHeaders.INVERTER_HEADERS)
        self._fill_example(sheet, SheetExamples.INVERTER_EXAMPLE)
        self.current_row += 1
